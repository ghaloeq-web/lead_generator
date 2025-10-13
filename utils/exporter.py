import os
import csv
import logging

logger = logging.getLogger(__name__)

def export_xlsx(leads, filename):
    """Export leads to Excel format"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        # Headers
        headers = ['Name', 'Title', 'Company', 'Phone', 'Email', 'Website', 'Industry', 'Location', 'Source']
        ws.append(headers)
        
        # Data
        for lead in leads:
            row = [
                lead.get('name', ''),
                lead.get('title', ''),
                lead.get('company', ''),
                lead.get('phone', ''),
                lead.get('email', ''),
                lead.get('website', ''),
                lead.get('industry', ''),
                lead.get('location', ''),
                lead.get('source', '')
            ]
            ws.append(row)
        
        wb.save(filename)
        logger.info(f"✅ Exported {len(leads)} leads to XLSX: {filename}")
        return True
    except ImportError:
        logger.warning("❌ openpyxl not available, falling back to CSV")
        export_csv(leads, filename.replace('.xlsx', '.csv'))
        return False
    except Exception as e:
        logger.error(f"❌ Error exporting to XLSX: {e}")
        export_csv(leads, filename.replace('.xlsx', '.csv'))
        return False

def export_pdf(leads, filename):
    """Export leads to PDF format"""
    try:
        from fpdf import FPDF
        
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        
        # Title
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Business Leads Report', 0, 1, 'C')
        pdf.ln(10)
        
        # Content
        pdf.set_font('Arial', '', 12)
        
        for i, lead in enumerate(leads, 1):
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, f'Lead #{i}: {lead.get("name", "N/A")}', 0, 1)
            pdf.set_font('Arial', '', 10)
            
            info = [
                f"Title: {lead.get('title', 'N/A')}",
                f"Company: {lead.get('company', 'N/A')}",
                f"Phone: {lead.get('phone', 'N/A')}",
                f"Email: {lead.get('email', 'N/A')}",
                f"Website: {lead.get('website', 'N/A')}",
                f"Industry: {lead.get('industry', 'N/A')}",
                f"Location: {lead.get('location', 'N/A')}",
                f"Source: {lead.get('source', 'N/A')}"
            ]
            
            for line in info:
                pdf.cell(0, 6, line, 0, 1)
            pdf.ln(5)
        
        pdf.output(filename)
        logger.info(f"✅ Exported {len(leads)} leads to PDF: {filename}")
        return True
    except ImportError:
        logger.warning("❌ fpdf not available, falling back to TXT")
        export_txt(leads, filename.replace('.pdf', '.txt'))
        return False
    except Exception as e:
        logger.error(f"❌ Error exporting to PDF: {e}")
        export_txt(leads, filename.replace('.pdf', '.txt'))
        return False

def export_vcf(leads, filename):
    """Export leads to vCard format"""
    try:
        import vobject
        
        with open(filename, 'w', encoding='utf-8') as f:
            for i, lead in enumerate(leads):
                vcard = vobject.vCard()
                
                # Add name
                if lead.get('name'):
                    names = lead['name'].split(' ', 1)
                    vcard.add('n')
                    if len(names) == 1:
                        vcard.n.value = vobject.vcard.Name(family=names[0], given='')
                    else:
                        vcard.n.value = vobject.vcard.Name(family=names[1], given=names[0])
                
                # Add formatted name
                if lead.get('name'):
                    vcard.add('fn')
                    vcard.fn.value = lead['name']
                
                # Add phone
                if lead.get('phone'):
                    vcard.add('tel')
                    vcard.tel.value = lead['phone']
                    vcard.tel.type_param = 'WORK'
                
                # Add email
                if lead.get('email'):
                    vcard.add('email')
                    vcard.email.value = lead['email']
                    vcard.email.type_param = 'WORK'
                
                # Add organization
                if lead.get('company'):
                    vcard.add('org')
                    vcard.org.value = [lead['company']]
                
                # Add title
                if lead.get('title'):
                    vcard.add('title')
                    vcard.title.value = lead['title']
                
                # Add note
                note_parts = []
                if lead.get('industry'):
                    note_parts.append(f"Industry: {lead['industry']}")
                if lead.get('source'):
                    note_parts.append(f"Source: {lead['source']}")
                
                if note_parts:
                    vcard.add('note')
                    vcard.note.value = ' | '.join(note_parts)
                
                f.write(vcard.serialize())
                f.write('\n')
        
        logger.info(f"✅ Exported {len(leads)} leads to VCF: {filename}")
        return True
    except ImportError:
        logger.warning("❌ vobject not available, falling back to CSV")
        export_csv(leads, filename.replace('.vcf', '.csv'))
        return False
    except Exception as e:
        logger.error(f"❌ Error exporting to VCF: {e}")
        export_csv(leads, filename.replace('.vcf', '.csv'))
        return False

def export_csv(leads, filename):
    """Export leads to CSV format"""
    try:
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Name', 'Title', 'Company', 'Phone', 'Email', 'Website', 'Industry', 'Location', 'Source'])
            
            for lead in leads:
                writer.writerow([
                    lead.get('name', ''),
                    lead.get('title', ''),
                    lead.get('company', ''),
                    lead.get('phone', ''),
                    lead.get('email', ''),
                    lead.get('website', ''),
                    lead.get('industry', ''),
                    lead.get('location', ''),
                    lead.get('source', '')
                ])
        logger.info(f"✅ Exported {len(leads)} leads to CSV: {filename}")
        return True
    except Exception as e:
        logger.error(f"❌ Error exporting to CSV: {e}")
        return False

def export_txt(leads, filename):
    """Export leads to simple text format"""
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("BUSINESS LEADS REPORT\n")
            f.write("=" * 60 + "\n\n")
            
            for i, lead in enumerate(leads, 1):
                f.write(f"LEAD #{i}\n")
                f.write("-" * 40 + "\n")
                f.write(f"Name: {lead.get('name', 'N/A')}\n")
                f.write(f"Title: {lead.get('title', 'N/A')}\n")
                f.write(f"Company: {lead.get('company', 'N/A')}\n")
                f.write(f"Phone: {lead.get('phone', 'N/A')}\n")
                f.write(f"Email: {lead.get('email', 'N/A')}\n")
                f.write(f"Website: {lead.get('website', 'N/A')}\n")
                f.write(f"Industry: {lead.get('industry', 'N/A')}\n")
                f.write(f"Location: {lead.get('location', 'N/A')}\n")
                f.write(f"Source: {lead.get('source', 'N/A')}\n")
                f.write("\n" + "=" * 60 + "\n\n")
        
        logger.info(f"✅ Exported {len(leads)} leads to TXT: {filename}")
        return True
    except Exception as e:
        logger.error(f"❌ Error exporting to TXT: {e}")
        return False

def export_data(leads, filename, format_type):
    """Main export function"""
    try:
        # Create directory if needed
        directory = os.path.dirname(filename)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        
        logger.info(f"📤 Exporting {len(leads)} leads as {format_type} to {filename}")
        
        # Export based on format
        if format_type == 'xlsx':
            return export_xlsx(leads, filename)
        elif format_type == 'pdf':
            return export_pdf(leads, filename)
        elif format_type == 'vcf':
            return export_vcf(leads, filename)
        elif format_type == 'csv':
            return export_csv(leads, filename)
        elif format_type == 'txt':
            return export_txt(leads, filename)
        else:
            logger.warning(f"❌ Unknown format {format_type}, defaulting to CSV")
            return export_csv(leads, filename.replace('.xlsx', '.csv').replace('.pdf', '.csv').replace('.vcf', '.csv'))
            
    except Exception as e:
        logger.error(f"❌ Error in export_data: {e}")
        # Last resort - try CSV
        try:
            csv_filename = 'leads_export.csv'
            return export_csv(leads, csv_filename)
        except Exception as final_error:
            logger.error(f"❌ Final export fallback failed: {final_error}")
            return False